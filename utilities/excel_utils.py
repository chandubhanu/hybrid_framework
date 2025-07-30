import openpyxl
from openpyxl.styles import PatternFill

def getRowCount(file,sheetName):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetName]
    return(sheet.max_row)

def getColumnCount(file,sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return (sheet.max_column)

def readData(file,sheetName,rowno,coloumnno):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return sheet.cell(rowno,coloumnno).value


def writeData(file,sheetName,row,coloumn,data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    sheet.cell(row,coloumn).value=data
    workbook.save(file)

def fillGreenCOlor(file,sheetName,roenum,coloumnno):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    greenFill=PatternFill(start_color='60b212',
                          end_color='60b212',
                          fill_type='solid')
    sheet.cell(roenum, coloumnno).fill = greenFill
    workbook.save(file)

def fillredColor(file,sheetName,roenum,coloumnno):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    redFill=PatternFill(start_color='ff0000',
                          end_color='ff0000',
                          fill_type='solid')
    sheet.cell(roenum, coloumnno).fill = redFill
    workbook.save(file)
